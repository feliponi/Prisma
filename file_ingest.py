"""File ingestion layer: turn arbitrary uploaded bytes into a `RawTable`.

Single responsibility: SOURCE-AGNOSTIC extraction of tabular structure
(headers + sample rows + row count) from CSV / XLS(X) / PDF bytes. This
module holds NO canonical-model semantics — it never maps to date/amount/etc.
and never computes transaction hashes. Mapping lives in mapping_ui.py and the
canonical transform lives in csv_mapper.py.

Runtime dependencies are imported lazily inside the readers so this module
stays importable without them:
    - openpyxl   -> Excel reading (`read_excel`, `list_excel_sheets`).
    - pdfplumber -> PDF table extraction (`read_pdf`).
CSV reading uses pandas + the standard library only.
"""

from __future__ import annotations

import csv as csv_lib
import io
import logging
from dataclasses import dataclass, field
from typing import Literal

import pandas as pd

logger = logging.getLogger(__name__)

# How many leading data rows to retain as a preview in every RawTable.
SAMPLE_ROW_LIMIT = 20

SourceType = Literal["csv", "xls", "pdf"]
PdfStrategy = Literal["lattice", "stream"]
Encoding = Literal["utf-8", "latin-1"]

# Candidate delimiters, most-specific first, used by the sniffer fallback.
_CANDIDATE_DELIMITERS = [";", ",", "\t", "|"]


@dataclass(frozen=True)
class CsvDialect:
    """Best-guess parsing parameters for a CSV, all user-overridable in the UI.

    Every field is a heuristic guess produced by `sniff_csv`; the mapping UI
    must expose each one for override before any parse is trusted.
    """

    delimiter: str
    encoding: Encoding
    decimal_separator: str
    thousands_separator: str
    quotechar: str = '"'
    has_header: bool = True


@dataclass(frozen=True)
class RawTable:
    """Source-agnostic tabular preview of an uploaded file.

    All sample values are raw strings exactly as extracted (no locale/number
    parsing here). `columns` are the detected header names in file order.
    """

    columns: list[str]
    sample_rows: list[dict[str, str]]
    row_count: int
    source_type: SourceType
    sheet_name: str | None = None
    pdf_strategy: str | None = None


@dataclass
class PdfExtractionResult:
    """Result wrapper for the fragile PDF path.

    PDF table extraction may fail or misalign; the reader NEVER raises on a
    bad table. Instead it returns this wrapper so the UI can surface warnings
    and require explicit user confirmation (or abort) before the extracted
    `raw_table` is trusted.
    """

    raw_table: RawTable | None
    warnings: list[str] = field(default_factory=list)
    needs_manual_review: bool = False


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------
def _decode_best(raw_bytes: bytes) -> tuple[str, Encoding]:
    """Decode bytes as utf-8, falling back to latin-1 (which never fails)."""
    try:
        return raw_bytes.decode("utf-8"), "utf-8"
    except UnicodeDecodeError:
        return raw_bytes.decode("latin-1"), "latin-1"


def _finalize_columns(header: list[str]) -> list[str]:
    """Clean header cells and disambiguate blanks/duplicates deterministically."""
    columns: list[str] = []
    seen: dict[str, int] = {}
    for i, raw in enumerate(header):
        name = ("" if raw is None else str(raw)).strip()
        if not name:
            name = f"column_{i + 1}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        columns.append(name)
    return columns


def _rows_to_sample(
    columns: list[str], data_rows: list[list[str]], max_rows: int | None
) -> list[dict[str, str]]:
    """Turn positional data rows into column-keyed dict rows (up to max_rows)."""
    limit = len(data_rows) if max_rows is None else max_rows
    sample: list[dict[str, str]] = []
    for row in data_rows[:limit]:
        record = {}
        for j, col in enumerate(columns):
            value = row[j] if j < len(row) else ""
            record[col] = "" if value is None else str(value)
        sample.append(record)
    return sample


def _infer_separators(text: str, delimiter: str) -> tuple[str, str]:
    """Infer (decimal, thousands) separators from numeric-looking cells.

    Recognizes BR-style "1.234,56" (decimal ",", thousands ".") and US-style
    "1,234.56" (decimal ".", thousands ","). Falls back to ("," , ".") — the
    Brazilian default this app is tuned for — when no evidence is found.
    """
    both_br = both_us = 0
    for raw_line in text.splitlines()[1:200]:  # skip header line
        for cell in raw_line.split(delimiter):
            token = cell.strip().strip('"')
            has_dot, has_comma = "." in token, "," in token
            if has_dot and has_comma:
                if token.rfind(",") > token.rfind("."):
                    both_br += 1
                else:
                    both_us += 1
    if both_us > both_br:
        return ".", ","
    return ",", "."


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------
def sniff_csv(raw_bytes: bytes) -> CsvDialect:
    """Detect a CSV's dialect by sampling: delimiter, encoding, and separators.

    Args:
        raw_bytes: The uploaded CSV file content.

    Returns:
        The best-guess `CsvDialect` (every field overridable in the UI).
    """
    text, encoding = _decode_best(raw_bytes)
    sample = "\n".join(text.splitlines()[:50])

    delimiter = ";"
    try:
        dialect = csv_lib.Sniffer().sniff(sample, delimiters="".join(_CANDIDATE_DELIMITERS))
        delimiter = dialect.delimiter
    except csv_lib.Error:
        # Fallback: pick the candidate with the highest, most consistent count.
        first_line = text.splitlines()[0] if text.splitlines() else ""
        counts = {d: first_line.count(d) for d in _CANDIDATE_DELIMITERS}
        delimiter = max(counts, key=counts.get) if any(counts.values()) else ","

    decimal_separator, thousands_separator = _infer_separators(text, delimiter)
    logger.info(
        "sniff_csv: delimiter=%r encoding=%s decimal=%r thousands=%r",
        delimiter, encoding, decimal_separator, thousands_separator,
    )
    return CsvDialect(
        delimiter=delimiter,
        encoding=encoding,
        decimal_separator=decimal_separator,
        thousands_separator=thousands_separator,
    )


def read_csv(raw_bytes: bytes, dialect: CsvDialect, max_rows: int | None = SAMPLE_ROW_LIMIT) -> RawTable:
    """Read CSV bytes into a `RawTable` using the (possibly user-edited) dialect.

    Args:
        raw_bytes: The uploaded CSV file content.
        dialect: The dialect to parse with (from `sniff_csv`, after overrides).
        max_rows: How many data rows to retain as the preview sample; pass
            None to retain every row (used by the full-import path).

    Returns:
        A `RawTable` with `source_type == "csv"`.

    Raises:
        UnicodeDecodeError: if `raw_bytes` cannot be decoded with the encoding.
        ValueError: if the content cannot be parsed as delimited text.
    """
    try:
        df = pd.read_csv(
            io.BytesIO(raw_bytes),
            delimiter=dialect.delimiter,
            dtype=str,
            keep_default_na=False,
            encoding=dialect.encoding,
            engine="python",
            quotechar=dialect.quotechar,
            header=0 if dialect.has_header else None,
        )
    except (pd.errors.ParserError, pd.errors.EmptyDataError) as exc:
        logger.error("read_csv failed: %s", exc)
        raise ValueError(f"Could not parse CSV: {exc}") from exc

    columns = _finalize_columns([str(c) for c in df.columns])
    data_rows = df.values.tolist()
    sample_rows = _rows_to_sample(columns, data_rows, max_rows)
    return RawTable(
        columns=columns,
        sample_rows=sample_rows,
        row_count=len(data_rows),
        source_type="csv",
    )


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
def list_excel_sheets(raw_bytes: bytes) -> list[str]:
    """List sheet names in an uploaded XLS/XLSX workbook (openpyxl engine).

    Args:
        raw_bytes: The uploaded Excel workbook content.

    Returns:
        Sheet names in workbook order.

    Raises:
        ValueError: if the bytes are not a readable Excel workbook.
    """
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise ValueError("openpyxl is required to read Excel files.") from exc
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
        return list(wb.sheetnames)
    except Exception as exc:  # noqa: BLE001 - any openpyxl failure -> clear error
        logger.error("list_excel_sheets failed: %s", exc)
        raise ValueError(f"Could not read Excel workbook: {exc}") from exc


def _excel_cell_to_str(value: object) -> str:
    """Render an openpyxl cell value as a raw string (None -> "")."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def read_excel(
    raw_bytes: bytes, sheet_name: str, header_row: int, max_rows: int | None = SAMPLE_ROW_LIMIT
) -> RawTable:
    """Read one Excel sheet into a `RawTable`, treating `header_row` as the header.

    Rows above `header_row` (e.g. an embedded title/subtitle row) are skipped;
    the row at `header_row` supplies `columns`; subsequent non-empty rows are data.

    Args:
        raw_bytes: The uploaded Excel workbook content.
        sheet_name: The sheet to read (from `list_excel_sheets`).
        header_row: Zero-based index of the row holding the column headers.
        max_rows: Preview sample size; None retains every data row.

    Returns:
        A `RawTable` with `source_type == "xls"` and `sheet_name` set.

    Raises:
        ValueError: if the sheet or header row cannot be read.
    """
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise ValueError("openpyxl is required to read Excel files.") from exc
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        ws = wb[sheet_name]
        all_rows = [[_excel_cell_to_str(c) for c in row] for row in ws.iter_rows(values_only=True)]
    except ValueError:
        raise
    except Exception as exc:  # noqa: BLE001 - any openpyxl failure -> clear error
        logger.error("read_excel failed: %s", exc)
        raise ValueError(f"Could not read Excel sheet: {exc}") from exc

    if header_row >= len(all_rows):
        raise ValueError(f"header_row={header_row} is beyond the sheet's {len(all_rows)} rows.")

    header = all_rows[header_row]
    columns = _finalize_columns(header)
    data_rows = [r for r in all_rows[header_row + 1:] if any(str(c).strip() for c in r)]
    sample_rows = _rows_to_sample(columns, data_rows, max_rows)
    return RawTable(
        columns=columns,
        sample_rows=sample_rows,
        row_count=len(data_rows),
        source_type="xls",
        sheet_name=sheet_name,
    )


# ---------------------------------------------------------------------------
# Bridge to the (unchanged) canonical transform
# ---------------------------------------------------------------------------
def raw_table_to_csv_bytes(raw: RawTable, delimiter: str, encoding: Encoding = "utf-8") -> bytes:
    """Serialize a `RawTable` back into delimited CSV bytes.

    This is the bridge that lets the UNCHANGED `csv_mapper.process_csv`
    (which parses delimited bytes) consume XLS- and PDF-sourced data too:
    ingest -> RawTable -> `raw_table_to_csv_bytes` -> `process_csv`. Fields are
    quoted as needed (csv.QUOTE_MINIMAL), so values containing the delimiter,
    quotes, or newlines round-trip safely.

    Args:
        raw: The ingested table to serialize (its `sample_rows` become the
            data lines — pass a full-read RawTable for a full import).
        delimiter: Field delimiter to emit (typically `profile.delimiter`).
        encoding: Output text encoding (must match `profile.encoding`).

    Returns:
        CSV-encoded bytes suitable for `csv_mapper.process_csv`.
    """
    buffer = io.StringIO()
    writer = csv_lib.writer(buffer, delimiter=delimiter, quoting=csv_lib.QUOTE_MINIMAL)
    writer.writerow(raw.columns)
    for record in raw.sample_rows:
        writer.writerow([record.get(col, "") for col in raw.columns])
    return buffer.getvalue().encode(encoding, errors="replace")


# ---------------------------------------------------------------------------
# PDF (fragile — never raises)
# ---------------------------------------------------------------------------
def read_pdf(
    raw_bytes: bytes, strategy: PdfStrategy, header_row: int, max_rows: int | None = SAMPLE_ROW_LIMIT
) -> PdfExtractionResult:
    """Attempt PDF table extraction (pdfplumber), never raising on a bad table.

    PDF is fragile and has no guaranteed columns. This function extracts tables
    with the given `strategy` ("lattice" for ruled tables, "stream" for
    whitespace-aligned tables), treats `header_row` within the best candidate
    table as the header, and — on any failure, missing table, ragged column
    counts, or a single-column blob — returns a result with a populated
    `warnings` list and `needs_manual_review=True` instead of raising.

    Args:
        raw_bytes: The uploaded PDF content.
        strategy: pdfplumber table-detection strategy.
        header_row: Zero-based header-row index within the detected table.
        max_rows: Preview sample size; None retains every data row.

    Returns:
        A `PdfExtractionResult`; `needs_manual_review` is True whenever the
        extraction is uncertain.
    """
    warnings: list[str] = []
    try:
        import pdfplumber
    except BaseException as exc:  # noqa: BLE001 - absent/broken native binding (may panic) must not crash
        logger.error("read_pdf: pdfplumber unavailable: %s", exc)
        return PdfExtractionResult(
            raw_table=None,
            warnings=["pdfplumber indisponível; extração de PDF não pôde ser executada."],
            needs_manual_review=True,
        )

    settings = (
        {"vertical_strategy": "lines", "horizontal_strategy": "lines"}
        if strategy == "lattice"
        else {"vertical_strategy": "text", "horizontal_strategy": "text"}
    )

    tables: list[list[list[str]]] = []
    try:
        with pdfplumber.open(io.BytesIO(raw_bytes)) as pdf:
            for page in pdf.pages:
                for extracted in page.extract_tables(table_settings=settings) or []:
                    cleaned = [["" if c is None else str(c).strip() for c in row] for row in extracted]
                    if cleaned:
                        tables.append(cleaned)
    except BaseException as exc:  # noqa: BLE001 - extraction (native code) must never raise
        logger.error("read_pdf: extraction failed: %s", exc)
        return PdfExtractionResult(
            raw_table=None,
            warnings=[f"Falha ao extrair tabela do PDF: {exc}"],
            needs_manual_review=True,
        )

    if not tables:
        return PdfExtractionResult(
            raw_table=None,
            warnings=["Nenhuma tabela detectada no PDF. Revise manualmente ou tente outra estratégia."],
            needs_manual_review=True,
        )

    table = max(tables, key=len)  # best candidate = most rows
    if header_row >= len(table):
        return PdfExtractionResult(
            raw_table=None,
            warnings=[f"Linha de cabeçalho {header_row} além da tabela extraída ({len(table)} linhas)."],
            needs_manual_review=True,
        )

    columns = _finalize_columns(table[header_row])
    data_rows = table[header_row + 1:]
    width = len(columns)
    ragged = any(len(r) != width for r in data_rows)
    if ragged:
        warnings.append("Colunas irregulares na extração do PDF — confira o alinhamento.")
    if width <= 1:
        warnings.append("Apenas uma coluna detectada — a extração provavelmente falhou.")

    sample_rows = _rows_to_sample(columns, data_rows, max_rows)
    raw_table = RawTable(
        columns=columns,
        sample_rows=sample_rows,
        row_count=len(data_rows),
        source_type="pdf",
        pdf_strategy=strategy,
    )
    # PDF extraction is never auto-trusted: always require an explicit confirm.
    needs_manual_review = True
    warnings.insert(0, "Extração de PDF é frágil; confirme os dados antes de prosseguir.")
    return PdfExtractionResult(
        raw_table=raw_table,
        warnings=warnings,
        needs_manual_review=needs_manual_review,
    )
